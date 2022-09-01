import datetime

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

from django.conf import settings
from celery import shared_task
from celery.utils.log import get_task_logger

from .models import InverterData, ZipReport, Location
from ..base.utils.timezone import localtime

logger = get_task_logger(__name__)


@shared_task(bind=True)
def generate_zip(extra_key=None, location_list=None, report_id=None, from_date=None, to_date=None):
    report_instance = ZipReport.objects.filter(pk=report_id).first()
    report_instance.status = "Generating"
    report_instance.save()
    for record in location_list:
        try:
            location = Location.objects.filter(pk=record).first()
            report_instance.location.add(location)
            context = None
            inverter_data = InverterData.objects.filter(device__location=location,
                                                        created_at__date__lte=to_date,
                                                        created_at__date__gte=from_date,
                                                        is_active=True).order_by('created_at').last()
            if inverter_data:
                context = {"total_energy": inverter_data.total_energy,
                           "daily_energy": inverter_data.daily_energy,
                           "op_active_power": inverter_data.op_active_power,
                           "specific_yields": inverter_data.specific_yields}
            wb = Workbook()
            sheet = wb['Sheet']
            wb.remove(sheet)
            ws1 = wb.create_sheet("Plant Summery")
            ws2 = wb.create_sheet("Plant Analysis")
            # ws3 = wb.create_sheet("Grid Downtime Analysis")
            # ws4 = wb.create_sheet("Inverter Summery ")
            # ws5 = wb.create_sheet("Alarm Analysis")
            # ws6 = wb.create_sheet("Help & Support")
            plant_summery_data = []
            plant_analysis_data = []

            if context and inverter_data:
                oap = float(inverter_data.op_active_power) if inverter_data.op_active_power else 0
                nominal_power = float(inverter_data.nominal_power) if inverter_data.nominal_power else 0
                irradiation = 0
                insolation = 0
                cuf = 0
                pr = 0
                if nominal_power != 0:
                    irradiation = (oap * 1361) / nominal_power
                    insolation = irradiation * 24
                    normal_irradiation = nominal_power * irradiation
                    cuf = (float(inverter_data.daily_energy) * 100) / (nominal_power * 24)
                    if normal_irradiation != 0:
                        pr = (oap * 1000 * 100) / normal_irradiation
                plant_summery_data = [
                    ['Plant Name', location.name],
                    ['Date', from_date,
                     to_date],
                    ['Description'],
                    ['Plant Capacity', location.capacity, "kWp"],
                    ['Plant Manager', location.manager],
                    ['Manager Phone', location.phone],
                    [''],
                    ['Daily Energy', context['daily_energy'] if 'daily_energy' in context else "--", "kWh"],
                    ['Output Active Power', context['op_active_power'] if 'op_active_power' in context else "--",
                     "kw"],
                    ['Specific Yield', context['specific_yields'] if 'specific_yields' in context else "--",
                     "(KWh/kwp)"],
                    ['CUF', cuf, "%"],
                    ['Performance Ratio', pr, "%"],
                    ['Total Energy', context['total_energy'] if 'total_energy' in context else "--", "kwh"],
                    ['Solar Insolation', insolation, "KWh/m2"],
                    ['Solar Irradiation', irradiation, "W/m2"],
                ]
                all_inverter_data = InverterData.objects.filter(device__location=location,
                                                                created_at__date__lte=to_date,
                                                                created_at__date__gte=from_date,
                                                                is_active=True)
                for inverter in all_inverter_data.iterator():
                    pr = 0
                    cuf = 0
                    irradiation = 0
                    insolation = 0
                    nominal_power = float(inverter.nominal_power) if inverter.nominal_power else 0
                    if nominal_power != 0:
                        irradiation = (oap * 1361) / nominal_power
                        insolation = irradiation * 24
                        normal_irradiation = nominal_power * irradiation
                        cuf = (float(inverter.daily_energy) * 100) / (nominal_power * 24)
                        if normal_irradiation != 0:
                            pr = (oap * 1000 * 100) / normal_irradiation
                    plant_analysis_data.append(
                        [localtime(inverter.created_at).replace(tzinfo=None), inverter.daily_energy,
                         inverter.op_active_power, inverter.specific_yields, cuf, pr, inverter.total_energy,
                         insolation, irradiation])
            else:
                plant_summery_data = [
                    ['Plant Name', location.name],
                    ['Date', from_date,
                     to_date],
                    ['Description'],
                    ['Plant Capacity', "", "kWp"],
                    ['Plant Manager'],
                    ['Manager Phone'],
                    ['']]
                plant_analysis_data = [['Error', "No data for the selected range"]]
            for row in plant_summery_data:
                ws1.append(row)
            ws2.append(["Timestamp", "Daily Energy [ KWh ]", "Output Active Power [ KW ]",
                        "Specific Yield [ KWh/kwp ]", "CUF [ % ]", "Performance Ratio [ % ]",
                        "Total Energy [ kwh ]", "Solar Insolation [ KWh/m2 ]", "Solar Irradiation [ W/m2 ]"])
            red_font = Font(bold=True, italic=True)
            for cell in ws2["1:1"]:
                cell.font = red_font
            for row in plant_analysis_data:
                ws2.append(row)
            Path(settings.MEDIA_ROOT + "/" + str(report_instance.id)).mkdir(parents=True, exist_ok=True)
            wb.save('{}/{}/{}.xlsx'.format(settings.MEDIA_ROOT, str(report_instance.id), location.name))
        except Exception as e:
            print(e)
            report_instance.status = "Error"
            report_instance.save()
            return None
    report_instance.status = "Success"
    report_instance.save()
    return None
